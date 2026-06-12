"""Products management router.

Provides CRUD endpoints for the products table, including:
- List all products with full specs (paginated, filterable)
- Update individual product fields (inline editing)
- Bulk update multiple products at once
"""

from datetime import datetime, timezone

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from loguru import logger
from pydantic import BaseModel
from sqlalchemy import func
from sqlmodel import select

from src.app.auth import get_current_user
from src.db.database import get_manual_db_session
from src.db.models.products import Product
from src.db.models.users import User

router = APIRouter(prefix="/products", tags=["products"])


# --- Schemas ---


class ProductResponse(BaseModel):
    id: str
    code: str
    name: str
    brand: str
    description: str
    data_rate: str
    fiber_type: str
    wavelength: str
    max_distance: str
    connector: str
    main_device: str
    category: str
    datasheet_path: str | None
    pdf_url: str | None
    raw_specs: str
    status: int
    created_at: str
    updated_at: str


class ProductListResponse(BaseModel):
    total: int
    products: list[ProductResponse]


class ProductUpdateRequest(BaseModel):
    code: str | None = None
    name: str | None = None
    brand: str | None = None
    description: str | None = None
    data_rate: str | None = None
    fiber_type: str | None = None
    wavelength: str | None = None
    max_distance: str | None = None
    connector: str | None = None
    main_device: str | None = None
    category: str | None = None
    raw_specs: str | None = None


class BulkUpdateItem(BaseModel):
    id: str
    changes: ProductUpdateRequest


class BulkUpdateRequest(BaseModel):
    items: list[BulkUpdateItem]


class BulkUpdateResponse(BaseModel):
    updated: int
    errors: list[str]


# --- Endpoints ---


class CreateProductRequest(BaseModel):
    code: str
    name: str = ""
    brand: str = ""
    description: str = ""
    data_rate: str = ""
    fiber_type: str = ""
    wavelength: str = ""
    max_distance: str = ""
    connector: str = ""
    main_device: str = "N/A"
    category: str = ""
    raw_specs: str = ""


@router.post("/", response_model=ProductResponse)
async def create_product(
    body: CreateProductRequest,
    current_user: User = Depends(get_current_user),
):
    """Create a new product manually."""
    if not body.code.strip():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Product code is required.")

    async with get_manual_db_session() as session:
        new_product = Product(
            code=body.code.strip(),
            name=body.name or body.code.strip(),
            brand=body.brand,
            description=body.description,
            data_rate=body.data_rate,
            fiber_type=body.fiber_type,
            wavelength=body.wavelength,
            max_distance=body.max_distance,
            connector=body.connector,
            main_device=body.main_device,
            category=body.category,
            raw_specs=body.raw_specs,
            status=1,
        )
        session.add(new_product)
        await session.flush()
        await session.refresh(new_product)

        return ProductResponse(
            id=new_product.id,
            code=new_product.code,
            name=new_product.name,
            brand=new_product.brand,
            description=new_product.description,
            data_rate=new_product.data_rate,
            fiber_type=new_product.fiber_type,
            wavelength=new_product.wavelength,
            max_distance=new_product.max_distance,
            connector=new_product.connector,
            main_device=new_product.main_device,
            category=new_product.category,
            datasheet_path=new_product.datasheet_path,
            pdf_url=new_product.pdf_url,
            raw_specs=new_product.raw_specs,
            status=new_product.status,
            created_at=new_product.created_at.isoformat() if new_product.created_at else "",
            updated_at=new_product.updated_at.isoformat() if new_product.updated_at else "",
        )


@router.get("/", response_model=ProductListResponse)
async def list_products(
    search: str | None = None,
    category: str | None = None,
    page: int = 1,
    page_size: int = 100,
    current_user: User = Depends(get_current_user),
):
    """List all products with full specs, supporting search and filtering."""
    async with get_manual_db_session() as session:
        query = select(Product).where(Product.status == 1)

        if category:
            query = query.where(Product.category == category)

        if search:
            search_term = f"%{search}%"
            query = query.where(
                Product.code.ilike(search_term) |
                Product.name.ilike(search_term) |
                Product.brand.ilike(search_term) |
                Product.description.ilike(search_term)
            )

        # Count total
        count_query = select(func.count()).select_from(query.subquery())
        total_result = await session.execute(count_query)
        total = total_result.scalar() or 0

        # Paginate
        query = query.offset((page - 1) * page_size).limit(page_size)
        query = query.order_by(Product.created_at.desc())

        result = await session.execute(query)
        products = result.scalars().all()

        return ProductListResponse(
            total=total,
            products=[
                ProductResponse(
                    id=p.id,
                    code=p.code,
                    name=p.name,
                    brand=p.brand,
                    description=p.description,
                    data_rate=p.data_rate,
                    fiber_type=p.fiber_type,
                    wavelength=p.wavelength,
                    max_distance=p.max_distance,
                    connector=p.connector,
                    main_device=p.main_device,
                    category=p.category,
                    datasheet_path=p.datasheet_path,
                    pdf_url=p.pdf_url,
                    raw_specs=p.raw_specs,
                    status=p.status,
                    created_at=p.created_at.isoformat() if p.created_at else "",
                    updated_at=p.updated_at.isoformat() if p.updated_at else "",
                )
                for p in products
            ],
        )


@router.patch("/{product_id}", response_model=ProductResponse)
async def update_product(
    product_id: str,
    body: ProductUpdateRequest,
    current_user: User = Depends(get_current_user),
):
    """Update a single product's fields (inline editing)."""
    async with get_manual_db_session() as session:
        result = await session.execute(
            select(Product).where(Product.id == product_id)
        )
        product = result.scalars().first()
        if not product:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Product not found")

        # Apply changes
        update_data = body.model_dump(exclude_none=True)
        for field, value in update_data.items():
            setattr(product, field, value)
        product.updated_at = datetime.now(timezone.utc)

        session.add(product)
        await session.commit()
        await session.refresh(product)

        return ProductResponse(
            id=product.id,
            code=product.code,
            name=product.name,
            brand=product.brand,
            description=product.description,
            data_rate=product.data_rate,
            fiber_type=product.fiber_type,
            wavelength=product.wavelength,
            max_distance=product.max_distance,
            connector=product.connector,
            main_device=product.main_device,
            category=product.category,
            datasheet_path=product.datasheet_path,
            pdf_url=product.pdf_url,
            raw_specs=product.raw_specs,
            status=product.status,
            created_at=product.created_at.isoformat() if product.created_at else "",
            updated_at=product.updated_at.isoformat() if product.updated_at else "",
        )


@router.post("/bulk-update", response_model=BulkUpdateResponse)
async def bulk_update_products(
    body: BulkUpdateRequest,
    current_user: User = Depends(get_current_user),
):
    """Bulk update multiple products at once (save all edited rows)."""
    updated = 0
    errors: list[str] = []

    async with get_manual_db_session() as session:
        for item in body.items:
            result = await session.execute(
                select(Product).where(Product.id == item.id)
            )
            product = result.scalars().first()
            if not product:
                errors.append(f"Product {item.id} not found")
                continue

            update_data = item.changes.model_dump(exclude_none=True)
            for field, value in update_data.items():
                setattr(product, field, value)
            product.updated_at = datetime.now(timezone.utc)
            session.add(product)
            updated += 1

        await session.commit()

    return BulkUpdateResponse(updated=updated, errors=errors)


@router.delete("/{product_id}")
async def delete_product(
    product_id: str,
    current_user: User = Depends(get_current_user),
):
    """Delete a product by ID."""
    async with get_manual_db_session() as session:
        result = await session.execute(
            select(Product).where(Product.id == product_id)
        )
        product = result.scalars().first()
        if not product:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Product not found")

        await session.delete(product)
        await session.commit()

    return {"message": f"Product '{product.code}' deleted."}


# --- Excel Sync ---


class ExcelFileResponse(BaseModel):
    id: str
    filename: str
    file_size: int
    total_rows: int
    created_at: str


class ExcelFileListResponse(BaseModel):
    total: int
    files: list[ExcelFileResponse]


class ExcelProductRefResponse(BaseModel):
    product_code: str
    description: str


class ExcelFileDetailResponse(BaseModel):
    id: str
    filename: str
    file_size: int
    total_rows: int
    created_at: str
    products: list[ExcelProductRefResponse]


class ExcelSyncResponse(BaseModel):
    total_rows_read: int
    matched: int
    updated: int
    not_found: list[str]
    message: str
    file_id: str


@router.get("/excel-files", response_model=ExcelFileListResponse)
async def list_excel_files(
    current_user: User = Depends(get_current_user),
):
    """List all uploaded Excel reference files."""
    from src.db.models.excel_refs import ExcelFile

    async with get_manual_db_session() as session:
        result = await session.execute(
            select(ExcelFile).order_by(ExcelFile.created_at.desc())
        )
        files = result.scalars().all()

        return ExcelFileListResponse(
            total=len(files),
            files=[
                ExcelFileResponse(
                    id=f.id,
                    filename=f.filename,
                    file_size=f.file_size,
                    total_rows=f.total_rows,
                    created_at=f.created_at.isoformat() if f.created_at else "",
                )
                for f in files
            ],
        )


@router.get("/excel-files/{file_id}", response_model=ExcelFileDetailResponse)
async def get_excel_file_detail(
    file_id: str,
    current_user: User = Depends(get_current_user),
):
    """Get detail of an Excel file including its product mappings."""
    from src.db.models.excel_refs import ExcelFile, ExcelProductRef

    async with get_manual_db_session() as session:
        result = await session.execute(
            select(ExcelFile).where(ExcelFile.id == file_id)
        )
        excel_file = result.scalars().first()
        if not excel_file:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File not found")

        refs_result = await session.execute(
            select(ExcelProductRef).where(ExcelProductRef.excel_file_id == file_id)
        )
        refs = refs_result.scalars().all()

        return ExcelFileDetailResponse(
            id=excel_file.id,
            filename=excel_file.filename,
            file_size=excel_file.file_size,
            total_rows=excel_file.total_rows,
            created_at=excel_file.created_at.isoformat() if excel_file.created_at else "",
            products=[
                ExcelProductRefResponse(product_code=r.product_code, description=r.description)
                for r in refs
            ],
        )


@router.delete("/excel-files/{file_id}")
async def delete_excel_file(
    file_id: str,
    current_user: User = Depends(get_current_user),
):
    """Delete an Excel file and its product mappings."""
    from sqlalchemy import delete as sa_delete
    from src.db.models.excel_refs import ExcelFile, ExcelProductRef

    async with get_manual_db_session() as session:
        result = await session.execute(
            select(ExcelFile).where(ExcelFile.id == file_id)
        )
        excel_file = result.scalars().first()
        if not excel_file:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File not found")

        # Delete refs first (foreign key constraint)
        await session.execute(
            sa_delete(ExcelProductRef).where(ExcelProductRef.excel_file_id == file_id)
        )
        await session.execute(
            sa_delete(ExcelFile).where(ExcelFile.id == file_id)
        )
        await session.commit()

    return {"message": f"File '{excel_file.filename}' and {excel_file.total_rows} mappings deleted."}


class ExcelProductRefUpdate(BaseModel):
    product_code: str
    description: str


class ExcelFileUpdateRequest(BaseModel):
    products: list[ExcelProductRefUpdate]


@router.patch("/excel-files/{file_id}")
async def update_excel_file_refs(
    file_id: str,
    body: ExcelFileUpdateRequest,
    current_user: User = Depends(get_current_user),
):
    """Update product mappings for an Excel file (inline editing)."""
    from src.db.models.excel_refs import ExcelFile, ExcelProductRef

    async with get_manual_db_session() as session:
        result = await session.execute(
            select(ExcelFile).where(ExcelFile.id == file_id)
        )
        excel_file = result.scalars().first()
        if not excel_file:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File not found")

        # Delete old refs and replace with new ones
        refs_result = await session.execute(
            select(ExcelProductRef).where(ExcelProductRef.excel_file_id == file_id)
        )
        for ref in refs_result.scalars().all():
            await session.delete(ref)
        await session.flush()

        # Insert updated refs
        for item in body.products:
            ref = ExcelProductRef(
                excel_file_id=file_id,
                product_code=item.product_code,
                description=item.description,
            )
            session.add(ref)

        excel_file.total_rows = len(body.products)
        session.add(excel_file)
        await session.commit()

    return {"message": f"Updated {len(body.products)} mappings for '{excel_file.filename}'."}


@router.post("/sync-excel", response_model=ExcelSyncResponse)
async def sync_from_excel(
    current_user: User = Depends(get_current_user),
):
    """Sync product descriptions from all uploaded Excel reference data.

    Reads all product_code → description mappings from the excel_product_refs table
    and bulk-updates matching products in the Products table.
    No file upload needed — uses data already stored via the Excel Refs tab.
    """
    from src.db.models.excel_refs import ExcelProductRef

    # Read all mappings from excel_product_refs
    pn_to_desc: dict[str, str] = {}
    async with get_manual_db_session() as session:
        refs_result = await session.execute(select(ExcelProductRef))
        all_refs = refs_result.scalars().all()

        for ref in all_refs:
            pn = ref.product_code.strip().upper()
            if pn and ref.description.strip():
                pn_to_desc[pn] = ref.description.strip()

    total_rows = len(pn_to_desc)

    if not pn_to_desc:
        return ExcelSyncResponse(
            total_rows_read=0, matched=0, updated=0, not_found=[],
            message="Không có dữ liệu Excel nào. Hãy upload file Excel ở tab Excel Refs trước.",
            file_id="",
        )

    # Bulk update matching products in DB
    matched = 0
    updated = 0
    not_found_codes: list[str] = []

    async with get_manual_db_session() as session:
        result = await session.execute(select(Product).where(Product.status == 1))
        all_products = result.scalars().all()
        product_map = {p.code.strip().upper(): p for p in all_products}

        for pn_upper, desc in pn_to_desc.items():
            product = product_map.get(pn_upper)
            if product:
                matched += 1
                if product.description != desc:
                    product.description = desc
                    product.updated_at = datetime.now(timezone.utc)
                    session.add(product)
                    updated += 1
            else:
                not_found_codes.append(pn_upper)

        await session.commit()

    return ExcelSyncResponse(
        total_rows_read=total_rows,
        matched=matched,
        updated=updated,
        not_found=not_found_codes[:50],
        message=f"Đồng bộ từ {total_rows} mã SP trong Excel Refs. Tìm thấy {matched} sản phẩm khớp, cập nhật {updated} mô tả.",
        file_id="",
    )


@router.post("/upload-excel", response_model=ExcelSyncResponse)
async def upload_excel(
    files: list[UploadFile] = File(...),
    current_user: User = Depends(get_current_user),
):
    """Upload Excel files, store mappings in DB, and sync product descriptions."""
    import io
    import openpyxl
    from src.db.models.excel_refs import ExcelFile, ExcelProductRef

    if not files:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No files provided.")

    total_rows = 0
    pn_to_desc: dict[str, str] = {}
    last_file_id = ""

    for file in files:
        filename = file.filename or ""
        if not filename.lower().endswith((".xlsx", ".xls")):
            continue

        content = await file.read()
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as e:
            logger.warning(f"Failed to read Excel file {filename}: {e}")
            continue

        # Extract P/N → Description mappings
        file_mappings: list[tuple[str, str]] = []

        for ws in wb.worksheets:
            pn_col = None
            desc_col = None
            header_row = None

            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), start=1):
                for cell in row:
                    val = str(cell.value or "").strip().lower()
                    if val in ("p/n", "pn", "product code", "mã sản phẩm", "part number"):
                        pn_col = cell.column - 1
                        header_row = row_idx
                    elif val in ("descriptions", "description", "mô tả", "desc"):
                        desc_col = cell.column - 1
                if pn_col is not None and desc_col is not None:
                    break

            if pn_col is None or desc_col is None:
                continue

            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if len(row) <= max(pn_col, desc_col):
                    continue
                pn = str(row[pn_col] or "").strip()
                desc = str(row[desc_col] or "").strip()
                if pn and desc:
                    file_mappings.append((pn, desc))
                    pn_to_desc[pn.upper()] = desc

        wb.close()

        # Save file record and mappings to DB
        if file_mappings:
            async with get_manual_db_session() as session:
                excel_file = ExcelFile(
                    filename=filename,
                    file_size=len(content),
                    total_rows=len(file_mappings),
                )
                session.add(excel_file)
                await session.flush()

                for pn, desc in file_mappings:
                    ref = ExcelProductRef(
                        excel_file_id=excel_file.id,
                        product_code=pn,
                        description=desc,
                    )
                    session.add(ref)

                await session.commit()
                last_file_id = excel_file.id
                total_rows += len(file_mappings)

    if not pn_to_desc:
        return ExcelSyncResponse(
            total_rows_read=0, matched=0, updated=0, not_found=[],
            message="No valid P/N + Description data found in the uploaded files.",
            file_id="",
        )

    # Bulk update matching products in DB
    matched = 0
    updated = 0
    not_found_codes: list[str] = []

    async with get_manual_db_session() as session:
        result = await session.execute(select(Product).where(Product.status == 1))
        all_products = result.scalars().all()
        product_map = {p.code.strip().upper(): p for p in all_products}

        for pn_upper, desc in pn_to_desc.items():
            product = product_map.get(pn_upper)
            if product:
                matched += 1
                if product.description != desc:
                    product.description = desc
                    product.updated_at = datetime.now(timezone.utc)
                    session.add(product)
                    updated += 1
            else:
                not_found_codes.append(pn_upper)

        await session.commit()

    return ExcelSyncResponse(
        total_rows_read=total_rows,
        matched=matched,
        updated=updated,
        not_found=not_found_codes[:50],
        message=f"Đã đọc {total_rows} dòng từ Excel. Tìm thấy {matched} sản phẩm khớp, cập nhật {updated} mô tả.",
        file_id=last_file_id,
    )


# --- Datasheet lookup ---


class DatasheetLookupRequest(BaseModel):
    product_codes: list[str]


class DatasheetLinkItem(BaseModel):
    product_code: str
    pdf_url: str | None
    datasheet_path: str | None


class DatasheetLookupResponse(BaseModel):
    """Response with deduplicated datasheet links for given product codes."""
    items: list[DatasheetLinkItem]
    unique_pdfs: list[str]


@router.post("/datasheets-lookup", response_model=DatasheetLookupResponse)
async def lookup_datasheets(
    body: DatasheetLookupRequest,
    current_user: User = Depends(get_current_user),
):
    """Look up datasheet PDF URLs for a list of product codes.

    Exact match only (case-insensitive). Returns per-product links and
    a deduplicated list of unique PDFs.
    """
    if not body.product_codes:
        return DatasheetLookupResponse(items=[], unique_pdfs=[])

    async with get_manual_db_session() as session:
        normalized = [c.strip().upper() for c in body.product_codes]
        result = await session.execute(
            select(Product).where(func.upper(Product.code).in_(normalized))
        )
        products = result.scalars().all()
        product_map = {p.code.strip().upper(): p for p in products}

    items: list[DatasheetLinkItem] = []
    seen_pdfs: set[str] = set()

    for code in body.product_codes:
        key = code.strip().upper()
        product = product_map.get(key)

        pdf_url = product.pdf_url if product else None
        datasheet_path = product.datasheet_path if product else None

        items.append(DatasheetLinkItem(
            product_code=code,
            pdf_url=pdf_url,
            datasheet_path=datasheet_path,
        ))

        if pdf_url:
            seen_pdfs.add(pdf_url)

    return DatasheetLookupResponse(
        items=items,
        unique_pdfs=sorted(seen_pdfs),
    )
