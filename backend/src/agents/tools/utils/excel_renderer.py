"""Excel renderer for quotation output.

Uses the MVC template (bảngBaoGia.xlsx) and fills in customer info + product rows.
Keeps all 13 original columns (A-M). Discount columns (J, K, L) are left empty.
"""

import io
from datetime import datetime
from pathlib import Path

import httpx
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from loguru import logger

from src.agents.tools.schemas import GenerateBomOutput, ProductInventoryStatus

_TEMPLATE_PATH = Path("data/bảngBaoGia.xlsx")
_DATA_START_ROW = 12
_TEMPLATE_DATA_ROWS = 3   # rows 12, 13, 14
_TEMPLATE_TOTALS_ROW = 15  # row 15

_CELL_FONT = Font(name="Times New Roman", size=11)
_CELL_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style(cell, align=None):
    cell.font = _CELL_FONT
    cell.alignment = align or _CELL_ALIGN
    cell.border = _THIN_BORDER


def _download_image(url: str) -> bytes | None:
    """Download image and convert to PNG bytes for Excel embedding."""
    if not url or not url.startswith("http"):
        return None
    try:
        with httpx.Client(timeout=10, follow_redirects=True) as client:
            resp = client.get(url)
            if resp.status_code != 200 or len(resp.content) > 500_000:
                return None
        from PIL import Image
        pil_img = Image.open(io.BytesIO(resp.content))
        buf = io.BytesIO()
        pil_img.save(buf, format="PNG")
        buf.seek(0)
        return buf.getvalue()
    except Exception as e:
        logger.debug(f"Image download/convert failed ({url}): {e}")
    return None


def _unmerge_from_row(ws, from_row: int):
    """Unmerge all merged cell ranges touching from_row or below."""
    to_remove = [m for m in ws.merged_cells.ranges if m.max_row >= from_row]
    for m in to_remove:
        ws.unmerge_cells(str(m))


def _clear_row(ws, row: int, max_col: int = 13):
    """Clear all cell values in a row."""
    for col in range(1, max_col + 1):
        ws.cell(row=row, column=col).value = None


def _write_product_row(ws, row: int, idx: int, item) -> float:
    """Write one product row. Returns line total."""
    _style(ws.cell(row=row, column=1, value=idx + 1))
    _style(ws.cell(row=row, column=2, value=item.product_code))

    name = item.product_name
    if item.product_code and item.product_code in name:
        name = name.replace(item.product_code, "").strip().strip("-").strip()
    _style(ws.cell(row=row, column=3, value=name), _CELL_ALIGN_LEFT)

    # D: Image
    ws.cell(row=row, column=4).border = _THIN_BORDER
    img_data = _download_image(item.image_url)
    if img_data:
        try:
            from openpyxl.drawing.image import Image as XlImage
            from openpyxl.utils.units import pixels_to_EMU
            from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor

            img = XlImage(io.BytesIO(img_data))
            img_size = 120  # pixels
            img.width = img_size
            img.height = img_size

            # Center image in cell D using TwoCellAnchor with offsets
            # Col D index = 3 (0-based), row index = row-1 (0-based)
            col_width_px = 290  # col D ~38 chars ≈ 290px
            row_height_px = max(65, (len(item.description or "") // 35 + 1) * 13)
            x_offset = max(0, (col_width_px - img_size) // 2)
            y_offset = max(0, (row_height_px - img_size) // 2)

            marker = AnchorMarker(col=3, colOff=pixels_to_EMU(x_offset), row=row - 1, rowOff=pixels_to_EMU(y_offset))
            marker2 = AnchorMarker(col=3, colOff=pixels_to_EMU(x_offset + img_size), row=row - 1, rowOff=pixels_to_EMU(y_offset + img_size))
            img.anchor = TwoCellAnchor(_from=marker, to=marker2)

            ws.add_image(img)
        except Exception as e:
            logger.warning(f"Image embed failed for {item.product_code}: {e}")

    _style(ws.cell(row=row, column=5, value=item.category))
    _style(ws.cell(row=row, column=6, value=item.description), _CELL_ALIGN_LEFT)
    _style(ws.cell(row=row, column=7, value=item.quantity))
    _style(ws.cell(row=row, column=8, value=item.unit or "cái"))

    c = ws.cell(row=row, column=9, value=item.unit_price)
    _style(c)
    c.number_format = '#,##0'

    # J, K, L: discount cols — empty with border
    for col in (10, 11, 12):
        ws.cell(row=row, column=col).border = _THIN_BORDER

    _style(ws.cell(row=row, column=13, value=item.notes or ""), _CELL_ALIGN_LEFT)

    # Row height: calculate based on description text wrapping in col F
    desc_text = item.description or ""
    chars_per_line = 35
    num_lines = max(1, len(desc_text) // chars_per_line + desc_text.count('\n') + 1)
    ws.row_dimensions[row].height = max(65, num_lines * 13)

    return item.unit_price * item.quantity


def render_bom_excel(
    bom: GenerateBomOutput,
    output_dir: Path,
    inventory_statuses: list[ProductInventoryStatus] | None = None,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    slug = bom.customer_name.lower().replace(" ", "_")[:30]
    filepath = output_dir / f"{timestamp}_{slug}.xlsx"

    if not _TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Template not found: {_TEMPLATE_PATH}")

    wb = load_workbook(str(_TEMPLATE_PATH))
    ws = wb.active
    num_items = len(bom.line_items)

    # Unmerge all from row 12 down (data + totals area)
    _unmerge_from_row(ws, _DATA_START_ROW)

    # Adjust rows: template has 3 data rows (12-14) + 1 totals row (15)
    # We need exactly num_items data rows + 1 totals row
    if num_items > _TEMPLATE_DATA_ROWS:
        # Insert extra rows BEFORE the totals row
        extra = num_items - _TEMPLATE_DATA_ROWS
        ws.insert_rows(_DATA_START_ROW + _TEMPLATE_DATA_ROWS, extra)
    elif num_items < _TEMPLATE_DATA_ROWS:
        # Delete unused placeholder rows
        rows_to_del = _TEMPLATE_DATA_ROWS - num_items
        ws.delete_rows(_DATA_START_ROW + num_items, rows_to_del)

    # Fill customer info
    now = datetime.now()
    ws["M1"] = now.strftime("%d-%m-%Y")
    ws["M2"] = f"BG{now.strftime('%y%m%d%H%M')}"
    ws["M3"] = f"NGÀY {now.day} THÁNG {now.month:02d} NĂM {now.year}"
    ws["C6"] = f"Ông/Bà: {bom.customer_name}"
    ws["C7"] = f"Email: {bom.customer_email}" if bom.customer_email else "Email:"
    ws["C8"] = f"Điện thoại: {bom.customer_phone}" if bom.customer_phone else "Điện thoại:"
    ws["C9"] = f"Địa chỉ: {bom.customer_address}" if bom.customer_address else "Địa chỉ:"

    # Write product rows
    grand_total = 0
    for idx, item in enumerate(bom.line_items):
        row = _DATA_START_ROW + idx
        _clear_row(ws, row)
        grand_total += _write_product_row(ws, row, idx, item)

    # Write totals row (right after last data row)
    totals_row = _DATA_START_ROW + num_items
    _clear_row(ws, totals_row)

    ws.merge_cells(start_row=totals_row, start_column=1, end_row=totals_row, end_column=8)
    lbl = ws.cell(row=totals_row, column=1, value="TỔNG GIÁ TRỊ ĐƠN HÀNG")
    lbl.font = Font(name="Times New Roman", size=11, bold=True)
    lbl.alignment = Alignment(horizontal="center", vertical="center")
    lbl.border = _THIN_BORDER

    tc = ws.cell(row=totals_row, column=9, value=grand_total)
    tc.font = Font(name="Times New Roman", size=11, bold=True)
    tc.alignment = _CELL_ALIGN
    tc.border = _THIN_BORDER
    tc.number_format = '#,##0'

    for col in range(10, 14):
        ws.cell(row=totals_row, column=col).border = _THIN_BORDER

    ws.row_dimensions[totals_row].height = 50

    # Style the Lưu ý/Note section below totals
    _NOTE_FONT = Font(name="Times New Roman", size=11, bold=True, italic=True)
    for r in range(totals_row + 1, ws.max_row + 1):
        for col in range(1, 14):
            cell = ws.cell(row=r, column=col)
            if cell.value is not None:
                cell.font = _NOTE_FONT

    wb.save(filepath)
    return filepath
