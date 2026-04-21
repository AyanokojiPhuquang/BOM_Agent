"""Import products from Excel file into nhanh_products.

Clears ALL existing products and inserts new ones from the Excel file.

Usage (inside container):
    uv run python -m scripts.import_products <path_to_xlsx>

Example:
    uv run python -m scripts.import_products data/products.xlsx
"""

import asyncio
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

from src.db.database import get_manual_db_session
from src.db.models.nhanh import NhanhProduct


def _norm(val) -> str:
    """Normalize Unicode to NFC and strip whitespace."""
    if val is None:
        return ""
    return unicodedata.normalize("NFC", str(val).strip())


def read_excel(filepath: str) -> list[dict]:
    """Read products from Excel. Expects columns: code, name, description, unit, image, base_price, sell_price, brand_name, category_name."""
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return []

    # Find header row (contains "code")
    header_idx = 0
    for i, row in enumerate(rows):
        if row and any(str(c).strip().lower() == "code" for c in row if c):
            header_idx = i
            break

    headers = [str(c).strip().lower() if c else f"col_{i}" for i, c in enumerate(rows[header_idx])]
    products = []

    for row in rows[header_idx + 1:]:
        if not row or not row[0]:
            continue
        data = dict(zip(headers, row))
        code = _norm(data.get("code", ""))
        if not code:
            continue

        products.append({
            "code": code,
            "name": _norm(data.get("name", "")),
            "description": _norm(data.get("description", "")),
            "unit": _norm(data.get("unit", "")),
            "image": _norm(data.get("image", "")),
            "base_price": float(data.get("base_price", 0) or 0),
            "sell_price": float(data.get("sell_price", 0) or 0),
            "brand": _norm(data.get("brand_name", "")),
            "category": _norm(data.get("category_name", "")),
        })

    wb.close()
    return products


async def import_products(filepath: str):
    products = read_excel(filepath)
    if not products:
        print(f"No products found in {filepath}")
        return

    print(f"Found {len(products)} products in Excel")

    async with get_manual_db_session() as session:
        # Clear all existing products
        from sqlalchemy import text
        await session.execute(text("DELETE FROM nhanh_products"))
        await session.commit()
        print("Cleared existing products")

        # Insert new products
        for i, p in enumerate(products):
            product = NhanhProduct(
                nhanh_id=100000 + i,
                name=p["name"],
                code=p["code"],
                price=p["base_price"],
                import_price=p["base_price"],
                sell_price=p["sell_price"],
                status=1,
                remain=100,
                available=100,
                image=p["image"],
                description=p["description"],
                unit=p["unit"],
                brand_name=p["brand"],
                category_name=p["category"],
                datasheet_path=None,
            )
            session.add(product)

        await session.commit()
        print(f"Imported {len(products)} products")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: uv run python -m scripts.import_products <path_to_xlsx>")
        sys.exit(1)
    asyncio.run(import_products(sys.argv[1]))
